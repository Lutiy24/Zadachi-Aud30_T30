from flask import Flask, request, jsonify
import openpyxl

app = Flask(__name__)

EXCEL_FILE = "currency_rates.xlsx"

def load_rates():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    rates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rates.append({
            "Currency1": row[0],
            "Currency2": row[1],
            "Rate": float(str(row[2]).replace(',', '.'))
        })
    return rates

@app.route("/currencies", methods=["GET"])
def get_currencies():
    rates = load_rates()
    currencies = set()
    for rate in rates:
        currencies.add(rate["Currency1"])
        currencies.add(rate["Currency2"])
    return jsonify([{"code": c} for c in sorted(currencies)])

@app.route("/rates", methods=["GET"])
def get_rates():
    return jsonify(load_rates())

@app.route("/convert", methods=["POST"])
def convert_currency():
    data = request.get_json()
    currency_from = data.get("currency_from")
    currency_to = data.get("currency_to")
    amount = float(data.get("amount"))

    for rate in load_rates():
        if rate["Currency1"] == currency_from and rate["Currency2"] == currency_to:
            converted = amount * rate["Rate"]
            return jsonify({"converted": round(converted, 2)})

    return jsonify({"error": "Rate not found"}), 404

@app.route("/add_currency", methods=["POST"])
def add_currency():
    data = request.get_json()
    currency1 = data.get("Currency1")
    currency2 = data.get("Currency2")
    rate = str(data.get("Rate")).replace(',', '.')

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([currency1, currency2, rate])
    wb.save(EXCEL_FILE)
    return jsonify({"status": "added"}), 201

@app.route("/delete_currency", methods=["DELETE"])
def delete_currency():
    data = request.get_json()
    currency1 = data.get("Currency1")
    currency2 = data.get("Currency2")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2))
    for i, row in enumerate(rows, start=2):
        if row[0].value == currency1 and row[1].value == currency2:
            ws.delete_rows(i)
            wb.save(EXCEL_FILE)
            return jsonify({"status": "deleted"})

    return jsonify({"error": "Rate not found"}), 404

if __name__ == "__main__":
    app.run(debug=True)
